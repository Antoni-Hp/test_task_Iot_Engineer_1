from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import collections

wb = Workbook()
data_sheet = wb.create_sheet("Dane z logów", 0)  # tworzenie arkuszy i ich umiejscowienie
chart_sheet = wb.create_sheet("Wykresy", 1)
i_heating, i_motion, i_temp_sensor, i_door, i_window, i_heat_switch, i_kit_switch, i_wifi = ([] for i in range(8))
issue_headers = ['issue_heating', 'issue_motion', 'issue_temperature_sensor', 'issue_door', 'issue_window',
                 'issue_heating_switch', 'issue_kitchen_switch', 'issue_wifi']
excel_file_name = "logs_analysis.xlsx"  # nazwa budowanego pliku excel, ewentualnie ze ścieżką


def upload_data(data, headers):
    issue_cat = 0
    last_cell = 1
    issues = []
    data_sheet.append(headers)
    for keys in data:
        for unique_headers in headers:  # listuje pojedyńcze unikatowe nagłówki
            if unique_headers not in keys:  # uzupełnia wartości dla brakujących kluczy
                keys[unique_headers] = 'NO_DATA'
        _add_data_to_table(headers, keys, issues)
    for i in [i_heating, i_motion, i_temp_sensor, i_door, i_window, i_heat_switch, i_kit_switch, i_wifi]:
        issues.append(len(i))
    tab = Table(displayName="Tabela1",
                ref=f"A1:{get_column_letter(data_sheet.max_column)}{data_sheet.max_row}")  # tworzy tabele o zakresie od A1 do ostatniej komórki z danymi
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True,
                           showColumnStripes=True)
    tab.tableStyleInfo = style
    data_sheet.add_table(tab)
    _chart_count_issues(issues, last_cell)
    for issues in [i_heating, i_motion, i_temp_sensor, i_door, i_window, i_heat_switch, i_kit_switch, i_wifi]:
        last_cell += 18
        last_cell = _chart_issue(issues, last_cell, issue_cat)
        issue_cat += 1
    try:
        wb.save(excel_file_name)  # zapis pliku excel
    except PermissionError:
        print(f"Nie można zapisać pliku {excel_file_name}, błąd uprawnień")


def _add_data_to_table(headers, data, issues):
    data_tmp = {}
    values_tmp = []
    for key in headers:
        data_tmp[key] = data[key]
    for value in data_tmp.values():
        values_tmp.append(value)
    data_sheet.append(values_tmp)
    _count_the_issues(data)


def _count_the_issues(keys):
    if keys[issue_headers[0]] != '' and keys[issue_headers[0]] != 'NO_DATA':
        i_heating.append(keys[issue_headers[0]].strip())
    elif keys[issue_headers[1]] != '' and keys[issue_headers[1]] != 'NO_DATA':
        i_motion.append(keys[issue_headers[1]].strip())
    elif keys[issue_headers[2]] != '' and keys[issue_headers[2]] != 'NO_DATA':
        i_temp_sensor.append(keys[issue_headers[2]].strip())
    elif keys[issue_headers[3]] != '' and keys[issue_headers[3]] != 'NO_DATA':
        i_door.append(keys[issue_headers[3]].strip())
    elif keys[issue_headers[4]] != '' and keys[issue_headers[4]] != 'NO_DATA':
        i_window.append(keys[issue_headers[4]].strip())
    elif keys[issue_headers[5]] != '' and keys[issue_headers[5]] != 'NO_DATA':
        i_heat_switch.append(keys[issue_headers[5]].strip())
    elif keys[issue_headers[6]] != '' and keys[issue_headers[6]] != 'NO_DATA':
        i_kit_switch.append(keys[issue_headers[6]].strip())
    elif keys[issue_headers[7]] != '' and keys[issue_headers[7]] != 'NO_DATA':
        i_wifi.append(keys[issue_headers[7]].strip())


def _chart_count_issues(issues, last_cell):
    for head in range(len(issue_headers)):
        chart_sheet.append([issue_headers[head], issues[head]])
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = f"Ilość błędów w okresie {data_sheet['P2'].value} - {data_sheet[f'P{data_sheet.max_row}'].value}"
    chart1.y_axis.title = 'Ilość wystąpień'
    chart1.x_axis.title = 'rodzaj błędów'
    chart1.width = (len(issue_headers) * 3.5)
    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=8, max_col=2)
    cats = Reference(chart_sheet, min_col=1, min_row=1, max_row=8)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart_sheet.add_chart(chart1, f"D{last_cell}")


def _chart_issue(issue_heating, last_cell, issue_cat):
    if chart_sheet.max_row < last_cell:
        variable = last_cell - chart_sheet.max_row
    else:
        variable = 1
        last_cell = chart_sheet.max_row + variable
    chart_sheet.insert_rows(chart_sheet.max_row, amount=variable)  # dodaje z pustych wierszy przed wczytaniem danych
    sort_issues = collections.Counter(issue_heating)  # sortuje błędy od największej ilości do najmniejszej
    char_range = chart_sheet.max_row
    for issue in sort_issues:
        chart_sheet.append([issue, sort_issues[issue]])  # wczytuje do excela rodzaj błędu i ilość jego wystąpień
    chart_heating = BarChart()
    chart_heating.type = "col"
    chart_heating.style = 10
    chart_heating.title = f"Rodzaj błędów w kolumnie {issue_headers[issue_cat]}"
    chart_heating.y_axis.title = 'Ilość wystąpień'
    chart_heating.x_axis.title = 'Nazwa błędu'
    chart_heating.width = (len(sort_issues) * 3.5)
    if len(sort_issues) < 4:
        chart_heating.width = 10
    if len(sort_issues) > 17:
        chart_heating.height = len(sort_issues) / 2
    data = Reference(chart_sheet, min_col=2, min_row=char_range, max_row=chart_sheet.max_row, max_col=2)
    cats = Reference(chart_sheet, min_col=1, min_row=char_range, max_row=chart_sheet.max_row)
    chart_heating.add_data(data, titles_from_data=False)
    chart_heating.set_categories(cats)
    chart_sheet.add_chart(chart_heating, f"D{last_cell}")
    return last_cell
